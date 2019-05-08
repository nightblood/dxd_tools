from openpyxl import Workbook
import openpyxl

# **************** 生成财报表格 *************************
#
# 资产负债表
dict_blance_sheet = {
	'现金': ['cash', 'saveFrm_enLgBalSht_cashCheck'],
	'短期投资': ['shortIvtmt', 'shortIvtmtCheck'],
	'资产跌价准备': ['pvsForAstcp', 'pvsForAstcpCheck'],
	'短期投资净额': ['shortIvtmtNetval', 'shortIvtmtNetvalCheck'],
	'应收票据': ['notesRcv', 'saveFrm_enLgBalSht_notesRcvCheck'],
	'应收股利': ['dividendRcv','saveFrm_enLgBalSht_dividendRcvCheck'],
	'应收利息': ['interestRcv','saveFrm_enLgBalSht_interestRcvCheck'],
	'应收账款': ['acctRcv','acctRcvCheck'],
	'减：坏账准备': ['pvForBad','pvForBadCheck'],
	'应收账款净额': ['acctRcvNetval','acctRcvNetvalCheck'],
	'预付账款': ['advMny','saveFrm_enLgBalSht_advMnyCheck'],
	'期货保证金': ['saveFrm_enLgBalSht_coverCost','saveFrm_enLgBalSht_coverCostCheck'],
	'应收补贴款': ['sbsRcv','saveFrm_enLgBalSht_sbsRcvCheck'],
	'应收出口退税': ['saveFrm_enLgBalSht_eptTrRcv','saveFrm_enLgBalSht_eptTrRcvCheck'],
	'其他应收款': ['othRcv','saveFrm_enLgBalSht_othRcvCheck'],
	'存货': ['stock','stockCheck'],
	'其中：原材料': ['rawMaterial','rawMaterialCheck'],
	'产成品': ['fnsPdt','fnsPdtCheck'],
	'减：存货跌价准备': ['sitCpPpt','sitCpPptCheck'],
	'存货净值': ['stockNetVal','stockNetValCheck'],
	'待摊费用': ['dfdPpdEps','saveFrm_enLgBalSht_dfdPpdEpsCheck'],
	'待处理流动资产净损失': ['ddpCaNlm','saveFrm_enLgBalSht_ddpCaNlmCheck'],
	'一年期长期债券投资': ['longBondsYear','saveFrm_enLgBalSht_longBondsYearCheck'],
	'其他流动资产': ['othCrtAst','saveFrm_enLgBalSht_othCrtAstCheck'],
	'流动资产合计': ['ttlCrtAst','ttlCrtAstCheck'],
	'长期投资': ['longIvtmt','longIvtmtCheck'],
	'其中：长期股权投资': ['longEquityIvtmt','longEquityIvtmtCheck'],
	'长期债券投资': ['longTermBonds','longTermBondsCheck'],
	'合并价差': ['incorpPriceDiff','incorpPriceDiffCheck'],
	'长期投资合计': ['ttlLongIvtmt','ttlLongIvtmtCheck'],
	'固定资产原价': ['fxdAst','fxdAstCheck'],
	'减：累计折旧': ['acmDpt','acmDptCheck'],
	'固定资产净值': ['fxdAstNetval','fxdAstNetvalCheck'],
	'减：固定资产减值准备': ['fxdAstDvlArg','fxdAstDvlArgCheck'],
	'固定资产净额': ['fxdAstNetmey','fxdAstNetmeyCheck'],
	'工程物资': ['prjMtr','saveFrm_enLgBalSht_prjMtrCheck'],
	'在建工程': ['cstInPrj','saveFrm_enLgBalSht_cstInPrjCheck'],
	'固定资产清理': ['dspOfFxdAst','saveFrm_enLgBalSht_dspOfFxdAstCheck'],
	'待处理固定资产净损失': ['ddpFxdAstNlm','saveFrm_enLgBalSht_ddpFxdAstNlmCheck'],
	'固定资产合计': ['ttlFxdAst','ttlFxdAstCheck'],
	'无形资产': ['enLgBalShtitgAst','enLgBalShtitgAstCheck'],
	'其中：无形资产土地使用权': ['enLgBalShtldOcpRt','enLgBalShtldOcpRtCheck'],
	'递延资产': ['dfdLsAst','dfdLsAstCheck'],
	'其中：递延固定资产资产修理': ['dfdFxdastRp','dfdFxdastRpCheck'],
	'递延固定资产改良支出': ['dfdFxdastMdpt','dfdFxdastMdptCheck'],
	'其他长期资产': ['enLgBalShtothLongAst','enLgBalShtothLongAstCheck'],
	'其中：其他长期资产特准储备物资': ['enLgBalShtothLongAstSumt','enLgBalShtothLongAstSumtCheck'],
	'无形及其他资产合计': ['ttlItgOthAst','ttlItgOthAstCheck'],
	'递延税项借项': ['dfdTxsDb','dfdTxsDbCheck'],
	'资产总计': ['ttlAst','ttlAstCheck'],
	'短期借款': ['shortLoans','saveFrm_enLgBalSht_shortLoansCheck'],
	'应付票据': ['notesPyb','saveFrm_enLgBalSht_notesPybCheck'],
	'应付账款': ['acctPyb','saveFrm_enLgBalSht_acctPybCheck'],
	'预收账款': ['advFromCst','saveFrm_enLgBalSht_advFromCstCheck'],
	'应付工资': ['salaryPyb','saveFrm_enLgBalSht_salaryPybCheck'],
	'应付福利费': ['pybWfFee','saveFrm_enLgBalSht_pybWfFeeCheck'],
	'应付股利': ['dvdsPyb','saveFrm_enLgBalSht_dvdsPybCheck'],
	'应交税金': ['txsPyb','saveFrm_enLgBalSht_txsPybCheck'],
	'其他应交款': ['othPybs','saveFrm_enLgBalSht_othPybsCheck'],
	'其他应付款': ['othPyb','saveFrm_enLgBalSht_othPybCheck'],
	'预提费用': ['ardEps','saveFrm_enLgBalSht_ardEpsCheck'],
	'预计负债': ['saveFrm_enLgBalSht_itdLbt','saveFrm_enLgBalSht_itdLbtCheck'],
	'一年内到期的长期负债': ['longLbtDwoy','saveFrm_enLgBalSht_longLbtDwoyCheck'],
	'其他流动负债': ['othCrtLbt','saveFrm_enLgBalSht_othCrtLbtCheck'],
	'流动负债合计': ['ttlCrtLbt','ttlCrtLbtCheck'],
	'长期借款': ['longLoans','saveFrm_enLgBalSht_longLoansCheck'],
	'应付债券': ['bondsPyb','saveFrm_enLgBalSht_bondsPybCheck'],
	'长期应付款': ['longAcctPyb','saveFrm_enLgBalSht_longAcctPybCheck'],
	'专项应付款': ['saveFrm_enLgBalSht_specialAcctPyb','saveFrm_enLgBalSht_specialAcctPybCheck'],
	'其他长期负债': ['enLgBalShtothLongLbt','enLgBalShtothLongLbtCheck'],
	'其中：特准储备基金': ['enLgBalShtrsvFd','enLgBalShtrsvFdCheck'],
	'长期负债合计': ['ttlLongLbt','ttlLongLbtCheck'],
	'递延税款贷项': ['dfdTxCrd','saveFrm_enLgBalSht_dfdTxCrdCheck'],
	'负债合计': ['ttlLbt','ttlLbtCheck'],
	'少数股东权益': ['mntStkhdIts','mntStkhdItsCheck'],
	'实收资本': ['pdInCpt','pdInCptCheck'],
	'国家资本': ['saveFrm_enLgBalSht_ctyCpt','saveFrm_enLgBalSht_ctyCptCheck'],
	'集体资本': ['saveFrm_enLgBalSht_kltCpt','saveFrm_enLgBalSht_kltCptCheck'],
	'法人资本': ['cprCpt','cprCptCheck'],
	'其中：国有法人资本': ['sttCprCpt','sttCprCptCheck'],
	'其中：集体法人资本': ['kltCprCpt','kltCprCptCheck'],
	'个人资本': ['saveFrm_enLgBalSht_indCpt','saveFrm_enLgBalSht_indCptCheck'],
	'外商资本': ['saveFrm_enLgBalSht_frCpt','saveFrm_enLgBalSht_frCptCheck'],
	'资本公积': ['cptSps','saveFrm_enLgBalSht_cptSpsCheck'],
	'盈余公积': ['prfSps','prfSpsCheck'],
	'其中：法定盈余公积': ['lgPfSps','lgPfSpsCheck'],
	'其中：公益金': ['pwf','pwfCheck'],
	'其中：补充流动资本': ['cmptCrtCpt','cmptCrtCptCheck'],
	'未确认的投资损失': ['saveFrm_enLgBalSht_nasdIvtmtLoss','saveFrm_enLgBalSht_nasdIvtmtLossCheck'],
	'未分配利润': ['udsPrf','saveFrm_enLgBalSht_udsPrfCheck'],
	'外币报表折算差额': ['saveFrm_enLgBalSht_tslRsv','saveFrm_enLgBalSht_tslRsvCheck'],
	'所有者权益合计': ['ttlEq','ttlEqCheck'],
	'负债和所有者权益总计': ['ttlLbtEq','ttlLbtEqCheck'],
}
# 损益表
dict_income_statement = {
	'一、主营业务收入': ['fir1','sec1','thr1'],
	'其中：出口产品销售收入': ['fir2','sec2','thr2'],
	'进口产品销售收入': ['fir3','sec3','thr3'],
	'减：折扣与折让': ['fir4','sec4','thr4'],
	'二、主营业务收入净额': ['fir5','sec5','thr5'],
	'减：主营业务成本': ['fir6','sec6','thr6'],
	'其中：出口产品销售成本': ['fir7','sec7','thr7'],
	'主营业务税金及附加': ['fir8','sec8','thr8'],
	'经营费用': ['fir9','sec9','thr9'],
	'其他费用': ['fir10','sec10','thr10'],
	'加：递延收益': ['fir11','sec11','thr11'],
	'代购代销收入': ['fir12','sec12','thr12'],
	'其他收入': ['fir13','sec13','thr13'],
	'三、主营业务利润': ['fir14','sec14','thr14'],
	'加：其他业务利润': ['fir15','sec15','thr15'],
	'减：营业费用': ['fir16','sec16','thr16'],
	'管理费用': ['fir17','sec17','thr17'],
	'财务费用': ['fir18','sec18','thr18'],
	'其他费用2': ['fir19','sec19','thr19'],
	'四、营业利润': ['fir20','sec20','thr20'],
	'加：投资收益': ['fir21','sec21','thr21'],
	'期货收益': ['fir22','sec22','thr22'],
	'补贴收入': ['fir23','sec23','thr23'],
	'其中：补贴前亏损的企业补贴收入': ['fir24','sec24','thr24'],
	'营业外收入': ['fir25','sec25','thr25'],
	'其中：处置固定资产净收益': ['fir26','sec26','thr26'],
	'非货币性交易收益': ['fir27','sec27','thr27'],
	'出售无形资产收益': ['fir28','sec28','thr28'],
	'罚款净收入': ['fir29','sec29','thr29'],
	'其他收入2': ['fir30','sec30','thr30'],
	'其中：用以前年度含量工资节余弥补利润': ['fir31','sec31','thr31'],
	'减：营业外支出': ['fir32','sec32','thr32'],
	'其中：处置固定资产净损失': ['fir33','sec33','thr33'],
	'债务重组损失': ['fir34','sec34','thr34'],
	'罚款支出': ['fir35','sec35','thr35'],
	'捐赠支出': ['fir36','sec36','thr36'],
	'其他支出': ['fir37','sec37','thr37'],
	'其中：结转的含量工资包干节余': ['fir38','sec38','thr38'],
	'加：以前年度损益调整': ['fir39','sec39','thr39'],
	'五、利润总额': ['fir40','sec40','thr40'],
	'减：所得税': ['fir41','sec41','thr41'],
	'少数股东损益': ['fir42','sec42','thr42'],
	'加：未确认的投资损失': ['fir43','sec43','thr43'],
	'六、净利润': ['fir44','sec44','thr44'],
	'加：年初未分配利润': ['fir45','sec45','thr45'],
	'盈余公积补亏': ['fir46','sec46','thr46'],
	'其他调整因素': ['fir47','sec47','thr47'],
	'七、可供分配的利润': ['fir48','sec48','thr48'],
	'减：单项留用的利润': ['fir49','sec49','thr49'],
	'补充流动资本': ['fir50','sec50','thr50'],
	'提取法定盈余公积': ['fir51','sec51','thr51'],
	'提取法定公益金': ['fir52','sec52','thr52'],
	'提取职工奖励及福利基金': ['fir53','sec53','thr53'],
	'提取储备基金': ['fir54','sec54','thr54'],
	'提取企业发展基金': ['fir55','sec55','thr55'],
	'利润归还投资': ['fir56','sec56','thr56'],
	'其他': ['fir57','sec57','thr57'],
	'八、可供投资者分配的利润': ['fir58','sec58','thr58'],
	'减：应付优先股股利': ['fir59','sec59','thr59'],
	'提取任意盈余公积': ['fir60','sec60','thr60'],
	'应付普通股股利': ['fir61','sec61','thr61'],
	'转作资本的普通股股利': ['fir62','sec62','thr62'],
	'其他2': ['fir63','sec63','thr63'],
	'九、未分配的利润': ['fir64','sec64','thr64'],
	'其中：应由以后年度税前利润弥补的亏损': ['frmIncomeStatementSave_enLgIcmStmt_icdDrPbtNy','frmIncomeStatementSave_enLgIcmStmt_icdDrPbtNyCyca','frmIncomeStatementSave_enLgIcmStmt_icdDrPbtNyCheck'],
}
# 现金流量表
dict_flow_statement = {
	'销售商品、提供劳务收到的现金': ['saveCashFlowFrm_enLgCashFlow_scOrOlCash','saveCashFlowFrm_enLgCashFlow_scOrOlCashCheck'],
	'收到的税费返还': ['saveCashFlowFrm_enLgCashFlow_rfOfTafr','saveCashFlowFrm_enLgCashFlow_rfOfTafrCheck'],
	'收到的其他与经营活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocrrToOa','saveCashFlowFrm_enLgCashFlow_ocrrToOaCheck'],
	'现金流入小计': ['cashOpIs','cashOpIsCheck'],
	'购买商品、接受劳务支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForCol','saveCashFlowFrm_enLgCashFlow_cpForColCheck'],
	'支付给职工以及为职工支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpToAndFe','saveCashFlowFrm_enLgCashFlow_cpToAndFeCheck'],
	'支付的各项税费': ['saveCashFlowFrm_enLgCashFlow_txAndFp','saveCashFlowFrm_enLgCashFlow_txAndFpCheck'],
	'支付的其它与经营活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocprToOa','saveCashFlowFrm_enLgCashFlow_ocprToOaCheck'],
	'现金流出小计': ['cashOpOs','cashOpOsCheck'],
	'经营活动产生的现金流量净额': ['cfgFromOaa','cfgFromOaaCheck'],
	'收回投资所收到的现金': ['saveCashFlowFrm_enLgCashFlow_cashFromIw','saveCashFlowFrm_enLgCashFlow_cashFromIwCheck'],
	'取得投资收益所收到的现金': ['saveCashFlowFrm_enLgCashFlow_cashFmIti','saveCashFlowFrm_enLgCashFlow_cashFmItiCheck'],
	'处置固定资产、无形资产和其他长期资产所收回的现金净额': ['saveCashFlowFrm_enLgCashFlow_ncFmDfaIaaola','saveCashFlowFrm_enLgCashFlow_ncFmDfaIaaolaCheck'],
	'收到的其他与投资活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocrrToIvt','saveCashFlowFrm_enLgCashFlow_ocrrToIvtCheck'],
	'现金流入小计': ['cashItIs','cashItIsCheck'],
	'购建固定资产、无形资产和其他长期资产所支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForBfaIaaoli','saveCashFlowFrm_enLgCashFlow_cpForBfaIaaoliCheck'],
	'投资所支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForIm','saveCashFlowFrm_enLgCashFlow_cpForImCheck'],
	'支付的其他与投资活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocprToIa','saveCashFlowFrm_enLgCashFlow_ocprToIaCheck'],
	'现金流出小计': ['cashItOs','cashItOsCheck'],
	'投资活动产生的现金流量净额': ['cfgFmIaa','cfgFmIaaCheck'],
	'吸收投资所收到的现金': ['saveCashFlowFrm_enLgCashFlow_cashFmAi','saveCashFlowFrm_enLgCashFlow_cashFmAiCheck'],
	'借款所收到的现金': ['saveCashFlowFrm_enLgCashFlow_cashRcvBrw','saveCashFlowFrm_enLgCashFlow_cashRcvBrwCheck'],
	'收到的其他与筹资活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocrrToFa','saveCashFlowFrm_enLgCashFlow_ocrrToFaCheck'],
	'现金流入小计': ['cashFnIs','cashFnIsCheck'],
	'偿还债务所支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForDb','saveCashFlowFrm_enLgCashFlow_cpForDbCheck'],
	'分配股利、利润或偿付利息所支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForDvPi','saveCashFlowFrm_enLgCashFlow_cpForDvPiCheck'],
	'支付的其他与筹资活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocprToFa','saveCashFlowFrm_enLgCashFlow_ocprToFaCheck'],
	'现金流出小计': ['cashFnOs','cashFnOsCheck'],
	'筹资活动产生的现金流量净额': ['cfFmFaa','cfFmFaaCheck'],
	'四、汇率变动对现金的影响': ['saveCashFlowFrm_enLgCashFlow_efeRateChgCash','saveCashFlowFrm_enLgCashFlow_efeRateChgCashCheck'],
	'五、现金及现金等价物净增加额': ['tempEnLgCashFlowniOfCaCe','enLgCashFlowniOfCaCeCheck'],
}
# 现金流量表附表
dict_flow_statement1 = {
	'净利润': ['enLgCashFlowSchedulenetProfit'],
	'加： 计提的资产减值准备': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_addPnOfAsIptOfAs'],
	'固定资产拆旧': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_dpnOfFdAs'],
	'无形资产摊销': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_itbAsWd'],
	'长期待摊费用摊销': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_ltPdEsRvn'],
	'待摊费用减少（减：增加）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_dcePdEs'],
	'预提费用增加（减：减少）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_iceAdEs'],
	'处置固定、无形和其他长期资产的损失（减：收益）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_hdFasIasLasLs'],
	'固定资产报废损失': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_faLosForRt'],
	'财务费用': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_fnlEps'],
	'投资损失（减：收益）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_ivtLos'],
	'递延税款贷项（减：借项）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_dfeTaxCd'],
	'存货的减少（减：增加）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_decOfSk'],
	'经营性应收项目的减少（减：增加）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_decOfBsRs'],
	'经营性应付项目的增加（减：减少）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_iceOfBsRs'],
	'其他': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_other'],
	'经营活动产生的现金流量净额': ['ncfFrombsAg'],
	'债务转为资本': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_dtItoCl'],
	'一年内到期的可转换公司债券': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_mwayCcb'],
	'融资租入固定资产': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_fgForFfas'],
	'其他': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_otherTwo'],
	'三、现金及现金等价物净增加情况': [],
	'现金的期末余额': ['cashCashEnicashForEdBe'],
	'减：现金的期初余额': ['cashCashEniLsCashForEyBe'],
	'加：现金等价物的期末余额': ['cashCashEniaddCashEqtEdBe'],
	'减：现金等价物的期初余额': ['cashCashEniLsCashEqtEyBe'],
	'现金及现金等价物净增加额': ['cashCashEni'],
}
tuple_dict = (dict_blance_sheet, dict_income_statement, dict_flow_statement, dict_flow_statement1)

titles = ['资产负债表', '损益表', '现金流量表', '现金流量表附表']

def get_blance_sheet_data_from_excel():
	print('\n正在读取资产负债表')


def get_business_report_data_from_excel(file):
	res = []
	print('\n正在从【%s】读财报数据...' % file)
	wb = openpyxl.load_workbook(file)

	for sheetname in wb.sheetnames:
		print('正在读取%s...' % sheetname)
		data_sheet = []
		sheet = wb[sheetname]
		if sheetname == '现金流量表附表':
			dict_item = dict_flow_statement1
			col_names = ('B')
		elif sheetname == '现金流量表':
			dict_item = dict_flow_statement
			col_names = ('B', 'C')
		elif sheetname == '资产负债表':
			dict_item = dict_blance_sheet
			col_names = ('B', 'C')
		else:  # 损益表
			dict_item = dict_income_statement
			col_names = ('B', 'C', 'D')
		index = 2
		while sheet['A' + str(index)].value is not None and sheet['A' + str(index)].value.strip() != '':
			print('=', sheet['A' + str(index)].value)
			for col_num in col_names:
				data_sheet.append(sheet[col_num + str(index)].value)
			index += 1
		tmp_dict = {}
		data_i = 0

		for key in (i for item in dict_item.values() for i in item):
			print('====', key)
			tmp_dict[key] = data_sheet[data_i]
			data_i += 1

		res.append(tmp_dict)

	return res


def create_excel():
	wb = Workbook()

	for idx in range(len(titles)):
		if idx == 0:
			ws_tmp = wb.active
		else:
			ws_tmp = wb.create_sheet()
		ws_tmp.title = titles[idx]
		ws_tmp.column_dimensions['A'].width = 30.0
		ws_tmp.column_dimensions['B'].width = 30.0
		ws_tmp.column_dimensions['C'].width = 30.0
		index_tmp = 2
		for key in tuple_dict[idx].keys():
			ws_tmp.row_dimensions[index_tmp].height = 30.0
			ws_tmp['A' + str(index_tmp)] = key
			index_tmp = index_tmp + 1
	wb.save(r'C:\Users\Administrator\Desktop\财报.xlsx')


if __name__ == '__main__':
	print(get_business_report_data_from_excel())
