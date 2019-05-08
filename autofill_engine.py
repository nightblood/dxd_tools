#!C:\Users\Administrator\AppData\Local\Programs\Python\Python37\
# enconding=utf8

from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import openpyxl
from functools import reduce
# from qt.mainwindow import Ui_MainWindow
import selenium.common.exceptions
# from inspect import isfunction
# from PyQt5.QtWidgets import QMessageBox
# import os

cert_types = ['101居民身份证', '103护照', '104军官证', '105士兵证', '106警官证', '111港澳居民来往内地通行证', '112台湾居民来往大陆通行证', '113外国人永久居留证']
dict_cert_types = {
    '101居民身份证': '101',
    '103护照': '103',
    '104军官证': '104',
    '105士兵证': '105',
    '106警官证': '106',
    '111港澳居民来往内地通行证': '111',
    '112台湾居民来往大陆通行证': '112',
    '113外国人永久居留证': '113'
}
dict_work_sub = [
    {
        '中国共产党中央委员会和地方各级党组织负责人': '001',
        '国家机关及其工作机构负责人': '002',
        '民主党派和社会团体及其工作机构负责': '003',
        '事业单位负责人': '004',
        '企业负责人': '005',
    },
    {
        '科学研究人员': '111',
        '工程技术人员': '113',
        '农业技术人员': '117',
        '飞机和船舶技术人员': '118',
        '卫生专业技术人员': '119',
        '经济业务人员': '121',
        '金融业务人员': '122',
        '法律专业人员': '123',
        '教学人员': '124',
        '文学艺术工作人员': '125',
        '体育工作人员': '126',
        '新闻出版、文化工作人员': '127',
        '宗教职业者': '128',
        '其他专业技术人员': '129',
    },
    {
        '行政办公人员': '331',
        '安全保卫和消防人员': '332',
        '邮政和电信业务人员': '333',
        '其他办事人员和有关人员': '339',
    },
    {
        '购销人员': '441',
        '仓储人员': '442',
        '餐饮服务人员': '443',
        '饭店,旅游及健身娱乐场所服务人员': '444',
        '运输服务人员': '448',
        '医疗卫生辅助服务人员': '446',
        '社会服务和居民生活服务人员': '447',
        '其他商业、服务业人员': '449',
    },
    {
        '种植业生产人员': '551',
        '林业生产及野生动植物保护人员': '552',
        '畜牧业生产人员': '553',
        '渔业生产人员': '554',
        '水利设施管理养护人员': '555',
        '其他农、林、牧、渔、水利业生产人员': '559',
    },
    {
        '勘测及矿物开采人员': '661',
        '金属冶炼、轧制人员': '662',
        '化工产品生产人员': '664',
        '机械制造加工人员': '666',
        '机电产品装配人员': '667',
        '机械设备修理人员': '671',
        '电力设备安装、运行、检修及供电人员': '672',
        '电子元器件与设备制造、装配、调试及维修人员': '673',
        '橡胶和塑料制品生产人员': '674',
        '纺织、针织、印染人员': '675',
        '裁剪、缝纫和皮革、毛皮制品加工制作人员': '676',
        '粮油、食品、饮料生产加工及饲料生产加工人员': '677',
        '烟草及其制品加工人员': '678',
        '药品生产人员': '679',
        '木材加工、人造板生产、木制品制作及制浆、造纸和纸制品生产加工人员': '681',
        '建筑材料生产加工人员': '682',
        '玻璃、陶瓷、搪瓷及其制品生产加工人员': '683',
        '广播影视制品制作、播放及文物保护作业人员': '684',
        '印刷人员': '685',
        '工艺、美术品制作人员': '686',
        '文化教育、体育用品制作人员': '687',
        '工程施工人员': '688',
        '运输设备操作人员及有关人员': '691',
        '环境监测与废物处理人员': '692',
        '检验、计量人员': '693',
        '其他生产、运输设备操作人员及有关人员': '699',
    },
    {
        '军人': 'XX0',
    },
    {
        '不便分类的其他从业人员': 'YY0',

    },
    {
        '未知': 'Z00',
    }
]
dict_work_fatcher = {
    '国家机关、党群组织、企业、事业单位负责人': '0',
    '专业技术人员': '1',
    '办事人员和有关人员': '3',
    '商业、服务业人员': '4',
    '农、林、牧、渔、水利业生产人员': '5',
    '生产、运输设备操作人员及有关人员': '6',
    '军人': 'X',
    '不便分类的其他从业人员': 'Y',
    '未知': 'Z'
}
dict_booklet_owner = {
    '是': '1',
    '否': '0'
}
dict_study = {
    '博士': '61:010',
    '硕士': '61:020',
    '大学本科': '61:030',
    '大专和专科学校': '61:040',
    '中等专业技术学校': '61:050',
    '技工学校': '61:060',
    '高中': '61:070',
    '初中': '61:080',
    '小学': '61:090',
    '文盲或半文盲': '61:100',
    '其他': '61:999',
}
dict_marriage = {
    '未婚': '10',
    '已婚': '20',
    '丧偶': '30',
    '离婚': '40',
    '未说明的婚姻状况': '90',
}
dict_politics = {
    '群众': '0',
    '中国共产党党员': '1',
    '中国共产党预备党员': '2',
    '中国共产主义青年团团员': '3',
    '中国国民党革命委员会会员': '4',
    '中国民主同盟盟员': '5',
    '中国民主建国会会员': '6',
    '中国民主促进会会员': '7',
    '中国农工民主党党员': '8',
    '中国致公党党员': '9',
    '九三学社社员': 'A',
    '台湾民主自治同盟盟员': 'B',
    '无党派民主人士': 'C',
}
dict_home = {
    '自置': '1',
    '按揭': '2',
    '亲属楼宇': '3',
    '集体宿舍': '4',
    '租房': '5',
    '共有住宅': '6',
    '其他': '7',
    '未知': '8',
}
dict_identity = {
    '国家公务员': '11',
    '专业技术人员': '13',
    '职员': '17',
    '企业管理人员': '21',
    '工人': '24',
    '农民': '27',
    '学生': '31',
    '现役军人': '37',
    '自由职业者': '51',
    '个体经营者': '54',
    '无业人员': '70',
    '退（离）休人员': '80',
    '其他': '90',
}
dict_cust_type = {
    '农户': '11',
    '非农户': '12',
    '非居民': '13',
}
dict_level = {
    '潜在': '1',
    '普通': '2',
    '普通贵宾': '3',
    '重要贵宾': '4',
}
dict_select = {
    'select_slcStudyExp': dict_study,  # 最高学历
    'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6R0tWbm52M0tiSXANCg': dict_marriage,  # 婚姻状况
    'select_addCustForm_cFc1RzhpTTBsbEt2cytzY2w3bkRMREpNb21pMzR5ZW4NCg': dict_politics,  # 政治面貌
    'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6UEtHZyt2emhnMGd3T1pmM1VOWS96QT0NCg': dict_booklet_owner,  # 户主标识
    'select_workFather': dict_work_fatcher,  # 职业
    'select_workSub': dict_work_sub,  # 职业
    'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TVJjMzhJTENqcE1UKytxYmg4V2JORT0NCg': dict_home,  # 居住情况
    'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TktmUG1GY2JkcksNCg': dict_identity,  # 个人身份
    'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TWpQc05HY3B5NTANCg': dict_cust_type,  # 对私客户类型
    'select_addCustForm_custBaseInfo_custLev': dict_level,  # 客户层次
}
# 资产负债表
dict_blance_sheet = {
	'现金': ['cash','saveFrm_enLgBalSht_cashCheck'],
	'短期投资': ['shortIvtmt','shortIvtmtCheck'],
	'资产跌价准备': ['pvsForAstcp','pvsForAstcpCheck'],
	'短期投资净额': ['shortIvtmtNetval','shortIvtmtNetvalCheck'],
	'应收票据': ['notesRcv','saveFrm_enLgBalSht_notesRcvCheck'],
	'应收股利': ['dividendRcv','saveFrm_enLgBalSht_dividendRcvCheck'],
	'应收利息': ['interestRcv','saveFrm_enLgBalSht_interestRcvCheck'],
	'应收账款': ['acctRcv','acctRcvCheck'],
	'减：坏账准备': ['pvForBad','pvForBadCheck'],
	'应收账款净额': ['acctRcvNetval','acctRcvNetvalCheck'],
	'预付账款': ['advMny','saveFrm_enLgBalSht_advMnyCheck'],
	'期货保证金': ['saveFrm_enLgBalSht_coverCost','saveFrm_enLgBalSht_coverCostCheck'],
	'应收补贴款': ['sbsRcv','saveFrm_enLgBalSht_sbsRcvCheck'],
	'应收出口退税': ['saveFrm_enLgBalSht_eptTrRcv','saveFrm_enLgBalSht_eptTrRcvCheck'],
	'其他应收款': ['othRcv','saveFrm_enLgBalSht_othRcvCheck'],
	'存货': ['stock','stockCheck'],
	'其中：原材料': ['rawMaterial','rawMaterialCheck'],
	'产成品': ['fnsPdt','fnsPdtCheck'],
	'减：存货跌价准备': ['sitCpPpt','sitCpPptCheck'],
	'存货净值': ['stockNetVal','stockNetValCheck'],
	'待摊费用': ['dfdPpdEps','saveFrm_enLgBalSht_dfdPpdEpsCheck'],
	'待处理流动资产净损失': ['ddpCaNlm','saveFrm_enLgBalSht_ddpCaNlmCheck'],
	'一年期长期债券投资': ['longBondsYear','saveFrm_enLgBalSht_longBondsYearCheck'],
	'其他流动资产': ['othCrtAst','saveFrm_enLgBalSht_othCrtAstCheck'],
	'流动资产合计': ['ttlCrtAst','ttlCrtAstCheck'],
	'长期投资': ['longIvtmt','longIvtmtCheck'],
	'其中：长期股权投资': ['longEquityIvtmt','longEquityIvtmtCheck'],
	'长期债券投资': ['longTermBonds','longTermBondsCheck'],
	'合并价差': ['incorpPriceDiff','incorpPriceDiffCheck'],
	'长期投资合计': ['ttlLongIvtmt','ttlLongIvtmtCheck'],
	'固定资产原价': ['fxdAst','fxdAstCheck'],
	'减：累计折旧': ['acmDpt','acmDptCheck'],
	'固定资产净值': ['fxdAstNetval','fxdAstNetvalCheck'],
	'减：固定资产减值准备': ['fxdAstDvlArg','fxdAstDvlArgCheck'],
	'固定资产净额': ['fxdAstNetmey','fxdAstNetmeyCheck'],
	'工程物资': ['prjMtr','saveFrm_enLgBalSht_prjMtrCheck'],
	'在建工程': ['cstInPrj','saveFrm_enLgBalSht_cstInPrjCheck'],
	'固定资产清理': ['dspOfFxdAst','saveFrm_enLgBalSht_dspOfFxdAstCheck'],
	'待处理固定资产净损失': ['ddpFxdAstNlm','saveFrm_enLgBalSht_ddpFxdAstNlmCheck'],
	'固定资产合计': ['ttlFxdAst','ttlFxdAstCheck'],
	'无形资产': ['enLgBalShtitgAst','enLgBalShtitgAstCheck'],
	'其中：无形资产土地使用权': ['enLgBalShtldOcpRt','enLgBalShtldOcpRtCheck'],
	'递延资产': ['dfdLsAst','dfdLsAstCheck'],
	'其中：递延固定资产资产修理': ['dfdFxdastRp','dfdFxdastRpCheck'],
	'递延固定资产改良支出': ['dfdFxdastMdpt','dfdFxdastMdptCheck'],
	'其他长期资产': ['enLgBalShtothLongAst','enLgBalShtothLongAstCheck'],
	'其中：其他长期资产特准储备物资': ['enLgBalShtothLongAstSumt','enLgBalShtothLongAstSumtCheck'],
	'无形及其他资产合计': ['ttlItgOthAst','ttlItgOthAstCheck'],
	'递延税项借项': ['dfdTxsDb','dfdTxsDbCheck'],
	'资产总计': ['ttlAst','ttlAstCheck'],
	'短期借款': ['shortLoans','saveFrm_enLgBalSht_shortLoansCheck'],
	'应付票据': ['notesPyb','saveFrm_enLgBalSht_notesPybCheck'],
	'应付账款': ['acctPyb','saveFrm_enLgBalSht_acctPybCheck'],
	'预收账款': ['advFromCst','saveFrm_enLgBalSht_advFromCstCheck'],
	'应付工资': ['salaryPyb','saveFrm_enLgBalSht_salaryPybCheck'],
	'应付福利费': ['pybWfFee','saveFrm_enLgBalSht_pybWfFeeCheck'],
	'应付股利': ['dvdsPyb','saveFrm_enLgBalSht_dvdsPybCheck'],
	'应交税金': ['txsPyb','saveFrm_enLgBalSht_txsPybCheck'],
	'其他应交款': ['othPybs','saveFrm_enLgBalSht_othPybsCheck'],
	'其他应付款': ['othPyb','saveFrm_enLgBalSht_othPybCheck'],
	'预提费用': ['ardEps','saveFrm_enLgBalSht_ardEpsCheck'],
	'预计负债': ['saveFrm_enLgBalSht_itdLbt','saveFrm_enLgBalSht_itdLbtCheck'],
	'一年内到期的长期负债': ['longLbtDwoy','saveFrm_enLgBalSht_longLbtDwoyCheck'],
	'其他流动负债': ['othCrtLbt','saveFrm_enLgBalSht_othCrtLbtCheck'],
	'流动负债合计': ['ttlCrtLbt','ttlCrtLbtCheck'],
	'长期借款': ['longLoans','saveFrm_enLgBalSht_longLoansCheck'],
	'应付债券': ['bondsPyb','saveFrm_enLgBalSht_bondsPybCheck'],
	'长期应付款': ['longAcctPyb','saveFrm_enLgBalSht_longAcctPybCheck'],
	'专项应付款': ['saveFrm_enLgBalSht_specialAcctPyb','saveFrm_enLgBalSht_specialAcctPybCheck'],
	'其他长期负债': ['enLgBalShtothLongLbt','enLgBalShtothLongLbtCheck'],
	'其中：特准储备基金': ['enLgBalShtrsvFd','enLgBalShtrsvFdCheck'],
	'长期负债合计': ['ttlLongLbt','ttlLongLbtCheck'],
	'递延税款贷项': ['dfdTxCrd','saveFrm_enLgBalSht_dfdTxCrdCheck'],
	'负债合计': ['ttlLbt','ttlLbtCheck'],
	'少数股东权益': ['mntStkhdIts','mntStkhdItsCheck'],
	'实收资本': ['pdInCpt','pdInCptCheck'],
	'国家资本': ['saveFrm_enLgBalSht_ctyCpt','saveFrm_enLgBalSht_ctyCptCheck'],
	'集体资本': ['saveFrm_enLgBalSht_kltCpt','saveFrm_enLgBalSht_kltCptCheck'],
	'法人资本': ['cprCpt','cprCptCheck'],
	'其中：国有法人资本': ['sttCprCpt','sttCprCptCheck'],
	'其中：集体法人资本': ['kltCprCpt','kltCprCptCheck'],
	'个人资本': ['saveFrm_enLgBalSht_indCpt','saveFrm_enLgBalSht_indCptCheck'],
	'外商资本': ['saveFrm_enLgBalSht_frCpt','saveFrm_enLgBalSht_frCptCheck'],
	'资本公积': ['cptSps','saveFrm_enLgBalSht_cptSpsCheck'],
	'盈余公积': ['prfSps','prfSpsCheck'],
	'其中：法定盈余公积': ['lgPfSps','lgPfSpsCheck'],
	'其中：公益金': ['pwf','pwfCheck'],
	'其中：补充流动资本': ['cmptCrtCpt','cmptCrtCptCheck'],
	'未确认的投资损失': ['saveFrm_enLgBalSht_nasdIvtmtLoss','saveFrm_enLgBalSht_nasdIvtmtLossCheck'],
	'未分配利润': ['udsPrf','saveFrm_enLgBalSht_udsPrfCheck'],
	'外币报表折算差额': ['saveFrm_enLgBalSht_tslRsv','saveFrm_enLgBalSht_tslRsvCheck'],
	'所有者权益合计': ['ttlEq','ttlEqCheck'],
	'负债和所有者权益总计': ['ttlLbtEq','ttlLbtEqCheck'],
}
# 损益表
dict_income_statement = {
	'一、主营业务收入': ['fir1','sec1','thr1'],
	'其中：出口产品销售收入': ['fir2','sec2','thr2'],
	'进口产品销售收入': ['fir3','sec3','thr3'],
	'减：折扣与折让': ['fir4','sec4','thr4'],
	'二、主营业务收入净额': ['fir5','sec5','thr5'],
	'减：主营业务成本': ['fir6','sec6','thr6'],
	'其中：出口产品销售成本': ['fir7','sec7','thr7'],
	'主营业务税金及附加': ['fir8','sec8','thr8'],
	'经营费用': ['fir9','sec9','thr9'],
	'其他费用': ['fir10','sec10','thr10'],
	'加：递延收益': ['fir11','sec11','thr11'],
	'代购代销收入': ['fir12','sec12','thr12'],
	'其他收入': ['fir13','sec13','thr13'],
	'三、主营业务利润': ['fir14','sec14','thr14'],
	'加：其他业务利润': ['fir15','sec15','thr15'],
	'减：营业费用': ['fir16','sec16','thr16'],
	'管理费用': ['fir17','sec17','thr17'],
	'财务费用': ['fir18','sec18','thr18'],
	'其他费用2': ['fir19','sec19','thr19'],
	'四、营业利润': ['fir20','sec20','thr20'],
	'加：投资收益': ['fir21','sec21','thr21'],
	'期货收益': ['fir22','sec22','thr22'],
	'补贴收入': ['fir23','sec23','thr23'],
	'其中：补贴前亏损的企业补贴收入': ['fir24','sec24','thr24'],
	'营业外收入': ['fir25','sec25','thr25'],
	'其中：处置固定资产净收益': ['fir26','sec26','thr26'],
	'非货币性交易收益': ['fir27','sec27','thr27'],
	'出售无形资产收益': ['fir28','sec28','thr28'],
	'罚款净收入': ['fir29','sec29','thr29'],
	'其他收入2': ['fir30','sec30','thr30'],
	'其中：用以前年度含量工资节余弥补利润': ['fir31','sec31','thr31'],
	'减：营业外支出': ['fir32','sec32','thr32'],
	'其中：处置固定资产净损失': ['fir33','sec33','thr33'],
	'债务重组损失': ['fir34','sec34','thr34'],
	'罚款支出': ['fir35','sec35','thr35'],
	'捐赠支出': ['fir36','sec36','thr36'],
	'其他支出': ['fir37','sec37','thr37'],
	'其中：结转的含量工资包干节余': ['fir38','sec38','thr38'],
	'加：以前年度损益调整': ['fir39','sec39','thr39'],
	'五、利润总额': ['fir40','sec40','thr40'],
	'减：所得税': ['fir41','sec41','thr41'],
	'少数股东损益': ['fir42','sec42','thr42'],
	'加：未确认的投资损失': ['fir43','sec43','thr43'],
	'六、净利润': ['fir44','sec44','thr44'],
	'加：年初未分配利润': ['fir45','sec45','thr45'],
	'盈余公积补亏': ['fir46','sec46','thr46'],
	'其他调整因素': ['fir47','sec47','thr47'],
	'七、可供分配的利润': ['fir48','sec48','thr48'],
	'减：单项留用的利润': ['fir49','sec49','thr49'],
	'补充流动资本': ['fir50','sec50','thr50'],
	'提取法定盈余公积': ['fir51','sec51','thr51'],
	'提取法定公益金': ['fir52','sec52','thr52'],
	'提取职工奖励及福利基金': ['fir53','sec53','thr53'],
	'提取储备基金': ['fir54','sec54','thr54'],
	'提取企业发展基金': ['fir55','sec55','thr55'],
	'利润归还投资': ['fir56','sec56','thr56'],
	'其他': ['fir57','sec57','thr57'],
	'八、可供投资者分配的利润': ['fir58','sec58','thr58'],
	'减：应付优先股股利': ['fir59','sec59','thr59'],
	'提取任意盈余公积': ['fir60','sec60','thr60'],
	'应付普通股股利': ['fir61','sec61','thr61'],
	'转作资本的普通股股利': ['fir62','sec62','thr62'],
	'其他2': ['fir63','sec63','thr63'],
	'九、未分配的利润': ['fir64','sec64','thr64'],
	'其中：应由以后年度税前利润弥补的亏损': ['frmIncomeStatementSave_enLgIcmStmt_icdDrPbtNy','frmIncomeStatementSave_enLgIcmStmt_icdDrPbtNyCyca','frmIncomeStatementSave_enLgIcmStmt_icdDrPbtNyCheck'],
}
# 现金流量表
dict_flow_statement = {
	'销售商品、提供劳务收到的现金': ['saveCashFlowFrm_enLgCashFlow_scOrOlCash','saveCashFlowFrm_enLgCashFlow_scOrOlCashCheck'],
	'收到的税费返还': ['saveCashFlowFrm_enLgCashFlow_rfOfTafr','saveCashFlowFrm_enLgCashFlow_rfOfTafrCheck'],
	'收到的其他与经营活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocrrToOa','saveCashFlowFrm_enLgCashFlow_ocrrToOaCheck'],
	'现金流入小计': ['cashOpIs','cashOpIsCheck'],
	'购买商品、接受劳务支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForCol','saveCashFlowFrm_enLgCashFlow_cpForColCheck'],
	'支付给职工以及为职工支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpToAndFe','saveCashFlowFrm_enLgCashFlow_cpToAndFeCheck'],
	'支付的各项税费': ['saveCashFlowFrm_enLgCashFlow_txAndFp','saveCashFlowFrm_enLgCashFlow_txAndFpCheck'],
	'支付的其它与经营活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocprToOa','saveCashFlowFrm_enLgCashFlow_ocprToOaCheck'],
	'现金流出小计': ['cashOpOs','cashOpOsCheck'],
	'经营活动产生的现金流量净额': ['cfgFromOaa','cfgFromOaaCheck'],
	'收回投资所收到的现金': ['saveCashFlowFrm_enLgCashFlow_cashFromIw','saveCashFlowFrm_enLgCashFlow_cashFromIwCheck'],
	'取得投资收益所收到的现金': ['saveCashFlowFrm_enLgCashFlow_cashFmIti','saveCashFlowFrm_enLgCashFlow_cashFmItiCheck'],
	'处置固定资产、无形资产和其他长期资产所收回的现金净额': ['saveCashFlowFrm_enLgCashFlow_ncFmDfaIaaola','saveCashFlowFrm_enLgCashFlow_ncFmDfaIaaolaCheck'],
	'收到的其他与投资活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocrrToIvt','saveCashFlowFrm_enLgCashFlow_ocrrToIvtCheck'],
	'现金流入小计': ['cashItIs','cashItIsCheck'],
	'购建固定资产、无形资产和其他长期资产所支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForBfaIaaoli','saveCashFlowFrm_enLgCashFlow_cpForBfaIaaoliCheck'],
	'投资所支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForIm','saveCashFlowFrm_enLgCashFlow_cpForImCheck'],
	'支付的其他与投资活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocprToIa','saveCashFlowFrm_enLgCashFlow_ocprToIaCheck'],
	'现金流出小计': ['cashItOs','cashItOsCheck'],
	'投资活动产生的现金流量净额': ['cfgFmIaa','cfgFmIaaCheck'],
	'吸收投资所收到的现金': ['saveCashFlowFrm_enLgCashFlow_cashFmAi','saveCashFlowFrm_enLgCashFlow_cashFmAiCheck'],
	'借款所收到的现金': ['saveCashFlowFrm_enLgCashFlow_cashRcvBrw','saveCashFlowFrm_enLgCashFlow_cashRcvBrwCheck'],
	'收到的其他与筹资活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocrrToFa','saveCashFlowFrm_enLgCashFlow_ocrrToFaCheck'],
	'现金流入小计': ['cashFnIs','cashFnIsCheck'],
	'偿还债务所支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForDb','saveCashFlowFrm_enLgCashFlow_cpForDbCheck'],
	'分配股利、利润或偿付利息所支付的现金': ['saveCashFlowFrm_enLgCashFlow_cpForDvPi','saveCashFlowFrm_enLgCashFlow_cpForDvPiCheck'],
	'支付的其他与筹资活动有关的现金': ['saveCashFlowFrm_enLgCashFlow_ocprToFa','saveCashFlowFrm_enLgCashFlow_ocprToFaCheck'],
	'现金流出小计': ['cashFnOs','cashFnOsCheck'],
	'筹资活动产生的现金流量净额': ['cfFmFaa','cfFmFaaCheck'],
	'四、汇率变动对现金的影响': ['saveCashFlowFrm_enLgCashFlow_efeRateChgCash','saveCashFlowFrm_enLgCashFlow_efeRateChgCashCheck'],
	'五、现金及现金等价物净增加额': ['tempEnLgCashFlowniOfCaCe','enLgCashFlowniOfCaCeCheck'],
}
# 现金流量表附表
dict_flow_statement1 = {
	'净利润': ['enLgCashFlowSchedulenetProfit'],
	'加： 计提的资产减值准备': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_addPnOfAsIptOfAs'],
	'固定资产拆旧': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_dpnOfFdAs'],
	'无形资产摊销': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_itbAsWd'],
	'长期待摊费用摊销': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_ltPdEsRvn'],
	'待摊费用减少（减：增加）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_dcePdEs'],
	'预提费用增加（减：减少）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_iceAdEs'],
	'处置固定、无形和其他长期资产的损失（减：收益）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_hdFasIasLasLs'],
	'固定资产报废损失': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_faLosForRt'],
	'财务费用': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_fnlEps'],
	'投资损失（减：收益）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_ivtLos'],
	'递延税款贷项（减：借项）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_dfeTaxCd'],
	'存货的减少（减：增加）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_decOfSk'],
	'经营性应收项目的减少（减：增加）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_decOfBsRs'],
	'经营性应付项目的增加（减：减少）': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_iceOfBsRs'],
	'其他': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_other'],
	'经营活动产生的现金流量净额': ['ncfFrombsAg'],
	'债务转为资本': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_dtItoCl'],
	'一年内到期的可转换公司债券': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_mwayCcb'],
	'融资租入固定资产': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_fgForFfas'],
	'其他': ['saveCashFlowScheduleFrm_enLgCashFlowSchedule_otherTwo'],
	'现金的期末余额': ['cashCashEnicashForEdBe'],
	'减：现金的期初余额': ['cashCashEniLsCashForEyBe'],
	'加：现金等价物的期末余额': ['cashCashEniaddCashEqtEdBe'],
	'减：现金等价物的期初余额': ['cashCashEniLsCashEqtEyBe'],
	'现金及现金等价物净增加额': ['cashCashEni'],
}

afEngin = ''


class Engine(object):


    def __init__(self):
        self.__drv = ''
        self.dict_func = {
            '1. 财报录入': self.func_company_business_report,
            '2. 新增普惠金融客户 ': self.func_add_person_customer,
            '3. 新增对公信贷客户 ': self.func_in_devloping,
            '9. 退出程序 exit': self.finish,
        }
        self.confirm_dialog = ''
        self.alert_dialog = ''
        self.window = ''
        # self.init_engine()
        self.__exit_flag = False

    @staticmethod
    def get_instance():
        global afEngin
        if afEngin is None or afEngin == '':
            afEngin = Engine()
        return afEngin

    def init_engine(self):
        # 加启动配置
        # self.openChrome()
        print('init_engine()...')
        self.openIe()
        self.login_granite()
        # time.sleep(2)
        # input('成功登录大信贷平台后点击【回车】继续。')

    # 前台开启浏览器模式
    def openChrome(self):
        # 加启动配置
        option = webdriver.ChromeOptions()
        option.add_argument('disable-infobars')
        # 打开chrome浏览器
        self.__drv = webdriver.Chrome(chrome_options=option)

    def openIe(self):
        # iedriver = 'D:\\app\\anaconda\\IEDriverServer.exe'  # iedriver路径
        # os.environ["webdriver.ie.driver"] = iedriver  # 设置环境变量
        # self.__drv = webdriver.Ie(iedriver)
        self.__drv = webdriver.Ie()

    # 授权操作
    def search_baidu(self):
        url = "http://www.baidu.com"
        self.__drv.get(url)
        # 找到输入框并输入查询内容
        elem = self.__drv.find_element_by_id("kw")
        elem.send_keys("selenium")
        # 提交表单
        self.__drv.find_element_by_xpath("//*[@id='su']").click()
        print('登录完毕！')

    def login_baidu(self):
        url = 'https://passport.baidu.com/v2/?login&tpl=mn&u=http%3A%2F%2Fwww.baidu.com%2F&sms=5'
        self.__drv.get(url)

        time.sleep(2)
        self.__drv.find_element_by_id('TANGRAM__PSP_3__footerULoginBtn').click()
        time.sleep(2)
        uname = self.__drv.find_element_by_id('TANGRAM__PSP_3__userName')
        upwd = self.__drv.find_element_by_id('TANGRAM__PSP_3__password')
        su = self.__drv.find_element_by_id('TANGRAM__PSP_3__submit')

        if uname is None:
            print('can not find username input element')
        else:
            print(uname)
        uname.send_keys('165374900@163.com')
        upwd.send_keys('zlf19891001')
        su.click()

    def login_github(self):
        self.__drv.get('https://github.com/login?return_to=%2Fjoin')
        time.sleep(2)
        self.__drv.find_element_by_id('login_field').send_keys('***')
        self.__drv.find_element_by_id('password').send_keys('***')
        commit = self.__drv.find_element(By.NAME, 'commit')

        commit.submit()

    def login_granite(self):
        self.__drv.get('http://154.233.5.1:8080/granite/sso/login')
        time.sleep(2)
        # self.__drv.find_element_by_id('username').send_keys('***')
        # self.__drv.find_element_by_id('password').send_keys('***')

    def auto_fill(self):
        elements = self.__drv.find_elements_by_class_name('index_function_p')
        for i in elements:
            print(i.get_attribute('title'))
            if i.get_attribute('title') == '对公信贷客户信息维护':
                print('找到【对公信贷客户信息维护】')
        # ele = driver.find_element_by_css_selector("a[text()='普惠金融客户信息维护']").click()
        # driver.find_element_by_link_text(u'普惠金融客户信息维护').click()
        # print(ele)
        # while True:
        #     try:
        #         s = input('输入\n')
        #         ele = driver.find_element_by_link_text(s)
        #         if ele is not None:
        #             ele.click()
        #         else:
        #             print('没有该元素')
        #     except Exception as e:
        #         print(e)
        # driver.close()
        print('end auto_fill....')

    def search_client(self):
        element = self.__drv.find_element_by_id('user_No')
        if element is not None:
            element.send_keys('8912313')
        else:
            print('no user_No id element!!')

    def get_company_report_data(self):
        return ''

    def func_company_business_report(self):
        try:
            self.__drv.find_element_by_link_text('对公信贷客户信息维护').click()
        except Exception as e:
            self.__drv.find_element_by_link_text('对公信贷客户信息查询').click()
        time.sleep(5)
        try:
            Select(self.__drv.find_element_by_id('sltSearch')).select_by_index(['客户号', '证件信息', '客户名称'].index('证件信息'))
            Select(self.__drv.find_element_by_id('addForm_certInfo_keyId_certTyp'))\
                .select_by_index(['201注册登记证', '202营业执照', '203组织机构代码证', '204机构信用代码证', '301外贸许可证',
                                  '302开户许可证', '303特种行业许可证', '304外汇经营许可证', '305金融许可证', '401国税登记证',
                                  '402地税登记证', '501批文', '502企业名称预先核准通知书', '511贷款卡', '999公司其他证件'].index('203组织机构代码证'))

            self.__drv.find_element_by_id('txtSearchValue').send_keys('L29327456')
            self.__drv.find_element_by_id('addForm_0').click()
            time.sleep(3)
            self.__drv.find_element_by_link_text('详细').click()
            time.sleep(3)
            self.__drv.find_element_by_link_text('财务信息').click()
            time.sleep(1)
            #打开资产负债链接
            btns = self.__drv.find_elements_by_class_name('ui-button')
            for btn in btns:
                # print(btn.get_attribute('onclick'))
                if btn.get_attribute('onclick') is not None and btn.get_attribute('onclick').__contains__('financialInfoMaintenance.html'):
                    btn.click()
                    break
            time.sleep(2)
            self.__drv.find_element_by_id('btnAdd').click()
            time.sleep(2)
            self.business_report_autofill_task()
            self.test_business_report_html()
        except Exception as e:
            print(e)

    def test_business_report_html(self):
        #############  1. 资产负债表 ############
        # forms = self.__drv.find_elements_by_class_name('ui-form-style-2')
        ############## 2. 损益表 #############
        # self.__drv.find_element_by_link_text('损益表').click()
        # time.sleep(1)
        # self.__drv.find_element_by_id('add').click()
        # time.sleep(2)
        ################# 3. 现金流量表 ####################
        # self.__drv.find_element_by_link_text('现金流量表').click()
        # time.sleep(1)
        # self.__drv.find_element_by_id('btnAddCashFlow').click()
        # time.sleep(2)
        ################# 4. 现金流量表附表 ####################
        self.__drv.find_element_by_link_text('现金流量表附表').click()
        time.sleep(1)
        self.__drv.find_element_by_id('btnAddCashFlowSchedule').click()
        time.sleep(2)

        forms = self.__drv.find_elements_by_class_name('ui-form-style-2')

        content = '{'

        try:
            for form in forms:
                rows = form.find_elements_by_tag_name('tr')
                for row in rows:
                    try:

                        content = content + ('\'' + row.find_elements_by_tag_name('th')[0].text + '\': ')
                    except Exception as e:
                        pass
                    cols = row.find_elements_by_tag_name('td')
                    content_col = '['
                    for col in cols:
                        try:
                            content_col = content_col + '\'' + col.find_element_by_class_name('ui-text').get_attribute(
                                'id') + '\','
                        except Exception as e:
                            pass
                    content_col = content_col.strip(',')
                    content = content + content_col + ']\n'
        except Exception as e:
            print(e)
        print(content + '}')

    def get_business_report_data_from_excel(self):
        res = []
        print('\n正在从【财报.xlsx】读财报数据...')
        wb = openpyxl.load_workbook('财报.xlsx')

        for sheetname in wb.sheetnames:
            data_sheet = []
            sheet = wb[sheetname]
            if sheetname == '现金流量表附表':
                dict_item = dict_flow_statement1
                col_names = ('B')
            elif sheetname == '现金流量表':
                dict_item = dict_flow_statement
                col_names = ('B', 'C')
            elif sheetname == '资产负债表':
                dict_item = dict_blance_sheet
                col_names = ('B', 'C')
            else:  # 损益表
                dict_item = dict_income_statement
                col_names = ('B', 'C', 'D')
            index = 2
            while sheet['A' + str(index)].value is not None and sheet['A' + str(index)].value.strip() != '':
                for col_num in col_names:
                    data_sheet.append(sheet[col_num + str(index)].value)
                index += 1
            tmp_dict = {}
            data_i = 0

            for key in (i for item in dict_item.values() for i in item):
                tmp_dict[key] = data_sheet[data_i]
                data_i += 1

            res.append(tmp_dict)

        return res

    def business_report_autofill_task(self):
        print('business_report_autofill_task()...')
        # data = self.get_company_report_data()
        # self.__drv.find_element_by_id('saveFrm_pubFncDocInfo_fncPrincipal').send_keys('') # 名字
        # self.__drv.find_element_by_id('saveFrm_pubFncDocInfo_tel').send_keys('') # 联系电话
        # 是否审计
        # self.__drv.find_element_by_id('saveFrm_pubFncDocInfo_rptAuditInd' + ('1' if data.is_audit else '0')).click()
        # 资产信息
        # self.__drv.find_element_by_id('cash').send_keys('') # 联系电话
        # self.__drv.find_element_by_id('enLgBalSht_cashCheck').send_keys('') # 联系电话

        # self.__drv.find_element_by_id('cash').send_keys('') # 联系电话
        # self.__drv.find_element_by_id('enLgBalSht_cashCheck').send_keys('') # 联系电话
        # TODO: 有部分字段需要点击
        data = self.get_business_report_data_from_excel()
        for key, value in data:
            self.__drv.find_element_by_id(key).send_keys(value)

    # @process_log_print
    def func_add_person_customer(self):
        data = self.get_data_from_excel()
        for cust in data:
            try:
                self.__drv.find_element_by_link_text('普惠金融客户信息维护').click()
            except Exception as e:
                self.__drv.find_element_by_link_text('普惠金融客户信息查询').click()

                # driver.get('http://154.233.5.1:8080/pumice/inFinaCustomer/getCustomerIndex.html?custBaseInfo.custFlg=1')
                # print(e)

            time.sleep(5)
            # self.check_elements_test()
            self.__drv.find_element_by_id('addButton').click()
            time.sleep(3)
            index = self.get_dict_index(dict_cert_types, cust['cert_type'].strip())
            # check_elements_test()
            Select(self.__drv.find_element_by_id('slcCertType')).select_by_index(index)  # 证件类型
            self.__drv.find_element_by_id('txtCertId').send_keys(cust['cert_id'])  # 证件号码
            self.__drv.find_element_by_id('txtInFi').send_keys(cust['addCustForm_cFc1RzhpTTBsbEluMUNKVUVWK1FOV3RVakI5OWtXM1MNCg'])  # 新增资金需求（万元）
            buttons = self.__drv.find_elements_by_class_name('ui-state-default')
            for btn in buttons:
                if btn.text == '提交':
                    btn.click()
                    break
            time.sleep(2)
            self.__drv.find_element_by_link_text('详细').click()
            time.sleep(2)
            sub_index = 0
            for key in cust.keys():
                try:
                    if cust.get(key) is None or cust.get(key).strip() == '':
                        continue
                    if key.__contains__('cert_'):  # 证件信息
                        continue
                    if key.__contains__('phone_'):  # 电话信息
                        continue
                    if key.__contains__('addr_'):  # 地址信息
                        continue
                    if key.__contains__('info_'):  # 头像采集
                        continue
                    if key.__contains__('select_'):  # 下拉选择框
                        s = dict_select.get(key)
                        print(key.replace('select_', ''), s)
                        if s == dict_work_sub:
                            Select(self.__drv.find_element_by_id(key.replace('select_', ''))).select_by_index(list(s[sub_index].keys()).index(cust.get(key)))
                            pass
                        else:
                            Select(self.__drv.find_element_by_id(key.replace('select_', ''))).select_by_index(list(s.keys()).index(cust.get(key)))
                            if key == 'select_workFather':
                                sub_index = list(s.keys()).index(cust.get(key))
                        continue
                    element = self.__drv.find_element_by_id(key)
                    print(element.get_attribute('value'))
                    if cust.get(key) != element.get_attribute('value'):
                        element.clear()
                        element.send_keys(cust.get(key))

                except Exception as e:
                    print(e)
                    return {'code': -1, 'msg': e}
            return {'code': 0}
            # while True:
            #     user_input = self.window.raw_confirm_dialog('自动填单完成，点击【保存信息】提交客户信息后，是否继续？')
            #
            #     if user_input == QMessageBox.Yes:
            #         break
            #     elif user_input == QMessageBox.No:
            #         self.exe_complite()
            #         return {'code': 0}

            # check_elements_test(driver)

    # def check_elements_test(self):
    #     while True:
    #         s = input('输入你想找的元素(格式：id/name/class value) ps. 输入exit退出\n')
    #         if s == 'exit':
    #             break
    #         a = s.split(' ')
    #
    #         if len(a) != 2:
    #             print('异常输入！！！')
    #             continue
    #         print('查找%s = %s 的元素。。。' % (a[0], a[1]))
    #         try:
    #             if a[0] == 'id':
    #                 ele = self.__drv.find_element_by_id(a[1])
    #                 print(ele.text, ele.get_attribute('value'))
    #                 # ele.click()
    #             elif a[0] == 'name':
    #                 ele = self.__drv.find_element_by_name(a[1])
    #                 print(ele.text, ele.get_attribute('value'))
    #                 # ele.click()
    #             elif a[0] == 'class':
    #                 ele = self.__drv.find_elements_by_class_name(a[1])
    #                 for e in ele:
    #                     print(e.text, e.get_attribute('value'))
    #                     e.click()
    #                 # ele = driver.find_element_by_class_name(a[1])
    #                 # print(ele.text)
    #                 # ele.click()
    #             else:
    #                 print('输入异常！！')
    #         except Exception as e:
    #             print('没有找到。。。', e)

    # 返回-1，没有在字典中找到。返回-2，查找的字段为空。返回-3，校验字典类型不是dict
    def get_dict_index(self, dict_data, target_key):
        print(target_key)
        if type(dict_data) != dict:
            return -3
        if target_key is None or len(target_key) == 0:
            return -2
        index = 0
        for key in dict_data.keys():
            if key.__contains__(target_key):
                break
            index = index + 1
        return index if index < len(dict_data) else -1

    def check_data(self, custs):
        check_fields = [
            {'id': 'cert_type', 'name': '证件类型', 'dict': dict_cert_types},
            {'id': 'select_slcStudyExp', 'name': '最高学历', 'dict': dict_study},
            {'id': 'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6R0tWbm52M0tiSXANCg', 'name': '婚姻状态', 'dict': dict_marriage},
            {'id': 'select_addCustForm_cFc1RzhpTTBsbEt2cytzY2w3bkRMREpNb21pMzR5ZW4NCg', 'name': '政治面貌',
             'dict': dict_politics},
            {'id': 'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6UEtHZyt2emhnMGd3T1pmM1VOWS96QT0NCg', 'name': '户主标识',
             'dict': dict_booklet_owner},
            {'id': 'select_workFather', 'name': '职业', 'dict': dict_work_fatcher},
            {'id': 'select_workSub', 'name': '职业', 'dict': dict_work_sub},
            {'id': 'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TVJjMzhJTENqcE1UKytxYmg4V2JORT0NCg', 'name': '居住情况',
             'dict': dict_home},
            {'id': 'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TktmUG1GY2JkcksNCg', 'name': '个人身份',
             'dict': dict_identity},
            {'id': 'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TWpQc05HY3B5NTANCg', 'name': '对私客户类型',
             'dict': dict_cust_type},
            {'id': 'select_addCustForm_custBaseInfo_custLev', 'name': '客户层次', 'dict': dict_level},
        ]
        cust_id = 0
        for cust in custs:
            cust_id = cust_id + 1
            print('\n校验第 %d 个客户数据是否合法...' % (cust_id))
            for field in check_fields:
                index = self.get_dict_index(field.get('dict'), cust.get(field.get('id')))
                if index == -1:
                    print('错误：在%s中没有找到指定类型 %s！！！\n' % (field.get('dict'), cust.get(field.get('id'))))
                    self.finish()
                if index == -2:
                    print('警告：没有输入%s 字段，请在详情页面自行修改！\n' % (field.get('name')))

    def get_data_from_excel(self):
        print('\n正在从【新增普惠金融客户.xlsx】读取客户信息...')
        wb = openpyxl.load_workbook('新增普惠金融客户.xlsx')
        # print(wb.sheetnames)
        sheet = wb[wb.sheetnames[0]]
        data = []
        index = 5
        while sheet['A' + str(index)].value is not None and sheet['A' + str(index)].value.strip() != '':
            item = {
                # 业务需求
                'addCustForm_cFc1RzhpTTBsbEluMUNKVUVWK1FOV3RVakI5OWtXM1MNCg': sheet['A' + str(index)].value,
                'addCustForm_cFc1RzhpTTBsbElYcWl4NEMyMXloTms3dTMyNnZWTDUNCg': sheet['B' + str(index)].value,
                'addCustForm_cFc1RzhpTTBsbEluMUNKVUVWK1FOU2NFWm9YZ2NmdUlYZ1p4eTZTenV2Yz0NCg': sheet['C' + str(index)].value,
                # 基础信息
                'txtCustName': sheet['D' + str(index)].value,
                'select_slcStudyExp': sheet['E' + str(index)].value,  # 最高学历
                'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6R0tWbm52M0tiSXANCg': sheet['F' + str(index)].value,  # 婚姻状况
                'select_addCustForm_cFc1RzhpTTBsbEt2cytzY2w3bkRMREpNb21pMzR5ZW4NCg': sheet['G' + str(index)].value,  # 政治面貌
                'addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6UEtHZyt2emhnMGdIRUp3MHRDWTlKZ2llU2hxK2RhK0t3PT0NCg': sheet['H' + str(index)].value,
                'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6UEtHZyt2emhnMGd3T1pmM1VOWS96QT0NCg': sheet['I' + str(index)].value,  # 户主标识
                'addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6Rm1FTkxFRG9YQkQNCg': sheet['J' + str(index)].value,
                'select_workFather': sheet['K' + str(index)].value,  # 职业
                'select_workSub': sheet['L' + str(index)].value,  # 职业
                'addCustForm_custBaseInfo_workRemark': sheet['M' + str(index)].value,
                'addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6QkJ4V3ltN3JoWkl2T0pVcHVHZkkxOD0NCg': sheet['N' + str(index)].value,
                'addCustForm_cFc1RzhpTTBsbEoxY0tBRFZUOU95dWxsTzBJb1NPeEcNCg': sheet['O' + str(index)].value,
                'addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6R3o5UFFEQ2VpZ2ENCg': sheet['P' + str(index)].value,
                'txtlaborPpl': sheet['Q' + str(index)].value,
                'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TVJjMzhJTENqcE1UKytxYmg4V2JORT0NCg': sheet['R' + str(index)].value,  # 居住情况
                'txtcustIndustry': sheet['S' + str(index)].value,
                'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TktmUG1GY2JkcksNCg': sheet['T' + str(index)].value,  # 个人身份
                'select_addCustForm_V2JURlRWVEk1QVhOTXNCUTdDQUF6TWpQc05HY3B5NTANCg': sheet['U' + str(index)].value,  # 对私客户类型
                'select_addCustForm_custBaseInfo_custLev': sheet['V' + str(index)].value,  # 客户层次
                'info_custimage': sheet['W' + str(index)].value,
                'info_custimagedesc': sheet['X' + str(index)].value,
                'picker_custflag': sheet['Y' + str(index)].value,  # 客户标识
                # 证件信息
                'cert_type': sheet['Z' + str(index)].value,  # 证件类型
                'cert_id': sheet['AA' + str(index)].value,
                'cert_name': sheet['AB' + str(index)].value,
                'cert_start': sheet['AC' + str(index)].value,
                'cert_end': sheet['AD' + str(index)].value,
                'cert_place': sheet['AE' + str(index)].value,
                'cert_dept': sheet['AF' + str(index)].value,
                'cert_permanent': sheet['AG' + str(index)].value,
                # 电话信息
                'phone_type': sheet['AH' + str(index)].value,
                'phone_no': sheet['AI' + str(index)].value,
                # 地址信息
                'addr_type': sheet['AJ' + str(index)].value,
                'addr_belong': sheet['AK' + str(index)].value,
                'addr_mail': sheet['AL' + str(index)].value,
                'addr_': sheet['AM' + str(index)].value,
                'addr_longtitude': sheet['AN' + str(index)].value,
                'addr_latitude': sheet['AO' + str(index)].value,
            }
            index = index + 1
            # print(item)
            data.append(item)
        wb.close()

        # 校验数据正确性
        self.check_data(data)

        print('读取客户信息完成...\n')
        return data

    def finish(self):
        if self.__drv is not None:
            self.__drv.close()
        # input('\n点击【回车】结束程序\n')
        print('*************** 程序结束 ***************')
        time.sleep(2)
        exit(0)

    def exe_complite(self):
        pass

    def func_in_devloping(self):
        return {'code': -1, 'msg': '\n功能正在开发中，敬请期待。。。\n'}

    def click_enter(self):
        try:
            while True:
                self.__drv.find_element_by_class_name('ui-button-text-only').click()
                time.sleep(1)
        except Exception as e:
            print('ui-button-text-only', e)
            pass
        try:
            # self.check_elements_test()
            self.__drv.switch_to.frame('ui-outer-center1')
        except Exception as e:
            print('ui-outer-center1', e)

    def pos_func(self):
        print('click pos...')

    def neg_func(self):
        print('click neg...')

    def click_exe_func(self, index):
        print('click_exe..----.', index)
        try:
            # if isfunction(self.confirm_dialog):
            # self.window.confirm_dialog('confirm dialog', self.pos_func, self.neg_func)
            # print(list(self.dict_func.values())[index])
            return list(self.dict_func.values())[index]()
            # return self.func_in_devloping()
        except Exception as e:
            print(e)
            return {'code': -1, 'msg': str(e)}

    def fill_blance_sheet(self, data):
        """资产负载表"""
        for key, value in data:
            self.__drv.find_element_by_id(key).send_keys(value)

    def fill_income_statement(self, data):
        pass

    def fill_flow_statement(self, data):
        pass

    def fill_flow_statement1(self, data):
        pass

    def auto_fill_all(self, data):
        time.sleep(5)

    def fill_form(self, data):
        for key, value in data.items():
            print(key, value)
            self.__drv.find_element_by_id(key).send_keys(value)

# # 方法主入口
# if __name__ == '__main__':
#     print('\n*********************** 大信贷自动化 *****************************\n')
#
#     # 加启动配置
#     openChrome()
#     login_granite()
#     time.sleep(2)
#     input('成功登录大信贷平台后点击【回车】继续。')
#
#     try:
#         __drv.find_element_by_class_name('ui-button-text-only').click()
#
#         time.sleep(2)
#         __drv.switch_to.frame('ui-outer-center1')
#         # check_elements_test(drv)
#
#     except selenium.common.exceptions.ElementNotVisibleException as e:
#         print(e)
#         pass
#
#
#     while True:
#
#         print('\n************ 功能列表 ***************\n')
#         for func_key in dict_func:
#             print(func_key)
#         print('\n**************************************\n')
#         user_input = input('输入序号并回车：\n')
#
#         for func_key in dict_func.keys():
#             if func_key.__contains__(user_input):
#                 dict_func.get(func_key)(func_key)

